from openpyxl import load_workbook
from datetime import datetime


debts_list = []
retake_date_list = []
retake_mark = ['н/у', 'н/з', 'н/я']
subject_list = ['иняз', 'инфа', 'дискрет']
retake_date_list = []
retake_date_dict = {}

retake_date = load_workbook('даты_пересдач.xlsx')   
sheet_retake_date = retake_date.active

debts = load_workbook('долги.xlsx')
sheet_debts = debts.active


def quantity_count_row(sheet):
    count_row = 0

    for i in range(1, 4000):
        data_row = sheet.cell(row=i, column=1).value
        if data_row is None:
            break
        count_row += 1
    return count_row


def quantity_count_column(sheet):
    count_column = 0

    for i in range(1, 4000):
        data_column = sheet.cell(row=2, column=i).value
        if data_column is None:
            break
        count_column += 1
    return count_column


debts_count_row = quantity_count_row(sheet_debts)
debts_count_column = quantity_count_column(sheet_debts)

for i in range(2, debts_count_row+1):
    debts_list.append([sheet_debts.cell(row=i, column=1).value])

for x in range(2, debts_count_row+1):
    for i in range(1, debts_count_column):
        score = sheet_debts.cell(row=x, column=i).value.lower()
        if score == 'н/у' or score == 'н/з' or score == 'н/я':
            debts_list[x-2].append(sheet_debts.cell(row=1, column=i).value)
#возвращается многомерный список с учениками и их долгами

retake_date_count_row = quantity_count_row(sheet_retake_date)
retake_date_count_column = quantity_count_column(sheet_retake_date)

for subject in subject_list:
    retake_date_list = []
    for i in range(2, retake_date_count_row + 1):
        if sheet_retake_date.cell(row=i, column=1).value == subject:
            retake_date_list.append(sheet_retake_date.cell(row=i, column=2).value)
    retake_date_dict[subject] = retake_date_list
    retake_date_dict[subject].sort()
#возвразает retake_date_dict в котором хранятся значения - пермет и ключи - даты пересдач

student = str(input('введите имя фаилию студента: '))
for debts_name in debts_list:
    if debts_name[0] == student:
        for debts_subject in debts_name[1:]:
            subject_dates = retake_date_dict.get(debts_subject, '')
            subject_retake_date = ", ".join([date.strftime("%d-%m-%Y") for date in subject_dates])
            print(f'предмет - {debts_subject} | даты - {subject_retake_date}')
